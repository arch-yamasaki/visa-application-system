#!/usr/bin/env python3
"""Create metadata inventory for visa test-case raw documents."""

from __future__ import annotations

import json
import unicodedata
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEST_ROOT = ROOT / "data" / "test_cases"
RAW_ROOT = TEST_ROOT / "raw" / "申請書類"


def rel(path: Path) -> str:
    return str(path.relative_to(ROOT))


def pdf_pages(path: Path) -> int | None:
    try:
        result = subprocess.run(
            ["pdfinfo", str(path)],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        return None
    for line in result.stdout.splitlines():
        if line.startswith("Pages:"):
            return int(line.split(":", 1)[1].strip())
    return None


def xlsx_sheets(path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    sheets = []
    for sheet in workbook.worksheets:
        sheets.append({"name": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column})
    workbook.close()
    return sheets


def doc_role(path: Path) -> str:
    normalized_path = unicodedata.normalize("NFC", str(path)).lower()
    name = unicodedata.normalize("NFC", path.name).lower()
    parent = unicodedata.normalize("NFC", path.parent.name)
    if "履歴書（申請には使わない）" in normalized_path:
        return "unused_resume"
    if "オファーレター" in name or "雇用条件通知書" in parent:
        return "employment_terms"
    if "会社書類" in name:
        return "company_documents"
    if "ヒアリング" in name or "coe書類" in name:
        return "intake_spreadsheet"
    if name.endswith(".msg"):
        return "email_context"
    if "添付しなかった" in name:
        return "not_attached_reference"
    if "申請書類" in name:
        return "submitted_application_bundle"
    if path.suffix.lower() == ".pdf":
        return "applicant_document_bundle"
    return "other"


def case_id(path: Path) -> str:
    parts = path.relative_to(RAW_ROOT).parts
    if not parts:
        return "unknown"
    top = parts[0]
    if "１回目" in top or "1回目" in top:
        return "gijinkoku_a_company_round1"
    if "2回目" in top or "２回目" in top:
        return "gijinkoku_a_company_round2_family_japan"
    return "unknown"


def residence_status(path: Path) -> str:
    return "engineer_humanities_international"


def record(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower().lstrip(".")
    item: dict[str, Any] = {
        "path": rel(path),
        "file_name": path.name,
        "extension": suffix,
        "size_bytes": path.stat().st_size,
        "case_id": case_id(path),
        "document_role": doc_role(path),
        "expected_residence_status": residence_status(path),
        "contains_personal_information": True,
        "use_as_input": "申請には使わない" not in str(path),
        "notes": "",
    }
    if suffix == "pdf":
        item["pdf_pages"] = pdf_pages(path)
    if suffix == "xlsx":
        item["sheets"] = xlsx_sheets(path)
    if item["document_role"] == "unused_resume":
        item["notes"] = "raw参考資料。原則として正解データ生成の入力から除外する。"
    if item["document_role"] == "not_attached_reference":
        item["notes"] = "申請添付外の参考資料。期待値との差分確認用。"
    return item


def build_catalog(records: list[dict[str, Any]]) -> dict[str, Any]:
    cases = []
    for cid in ["gijinkoku_a_company_round1", "gijinkoku_a_company_round2_family_japan"]:
        case_records = [item for item in records if item["case_id"] == cid]
        roles: dict[str, int] = {}
        for item in case_records:
            roles[item["document_role"]] = roles.get(item["document_role"], 0) + 1
        cases.append(
            {
                "case_id": cid,
                "display_name": "A社 1回目申請" if cid.endswith("round1") else "A社 2回目申請 家族在住パターン",
                "target_status": "engineer_humanities_international",
                "application_type": "certificate_of_eligibility",
                "raw_source_dir": rel(RAW_ROOT / ("A社（１回目申請）" if cid.endswith("round1") else "A社（2回目申請）家族在住パターン")),
                "classification": {
                    "document_count": len(case_records),
                    "roles": roles,
                },
                "fixture_dir": f"data/test_cases/fixtures/{cid}",
                "golden_status": "skeleton_created",
            }
        )
    return {
        "schema_version": "0.1.0",
        "description": "Raw application-document catalog for AI extraction tests.",
        "raw_root": rel(RAW_ROOT),
        "cases": cases,
        "documents": records,
    }


def write_fixture(case: dict[str, Any], records: list[dict[str, Any]]) -> None:
    fixture_dir = TEST_ROOT / "fixtures" / case["case_id"]
    input_dir = fixture_dir / "input"
    expected_dir = fixture_dir / "expected"
    generated_dir = fixture_dir / "generated"
    input_dir.mkdir(parents=True, exist_ok=True)
    expected_dir.mkdir(parents=True, exist_ok=True)
    generated_dir.mkdir(parents=True, exist_ok=True)

    (fixture_dir / "scenario.json").write_text(
        json.dumps(
            {
                "schema_version": "0.1.0",
                "case_id": case["case_id"],
                "display_name": case["display_name"],
                "target_status": case["target_status"],
                "application_type": case["application_type"],
                "tags": [
                    "mvp",
                    "engineer_humanities_international",
                    "certificate_of_eligibility",
                    "raw_documents",
                ]
                + (["family_in_japan"] if "family" in case["case_id"] else ["round1"]),
                "primary_checks": [
                    "document_classification",
                    "case_data_extraction",
                    "form_mapping",
                    "gijinkoku_fit",
                    "missing_items",
                ],
                "input_documents": "input/input_documents.json",
                "expected_case_data": "expected/case_data.golden.json",
                "expected_application_data": "expected/application_data.golden.json",
                "expected_review": "expected/review.golden.json",
                "generated_dir": "generated",
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )

    case_records = [item for item in records if item["case_id"] == case["case_id"]]
    (input_dir / "input_documents.json").write_text(
        json.dumps(
            {
                "schema_version": "0.1.0",
                "case_id": case["case_id"],
                "target_status": case["target_status"],
                "application_type": case["application_type"],
                "documents": case_records,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )

    (expected_dir / "case_data.golden.json").write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "golden_status": "pending_manual_completion",
                "case": {
                    "case_id": case["case_id"],
                    "application_type": case["application_type"],
                    "target_status": case["target_status"],
                    "workflow_state": "golden_pending",
                    "intake_channel": "recruiting_agency",
                    "source_organization": "A社",
                },
                "expected_extraction_scope": {
                    "must_extract": [
                        "applicant",
                        "passport",
                        "immigration_history",
                        "family",
                        "education",
                        "employment_history",
                        "employer",
                        "application.activity_details",
                        "supporting_documents",
                    ],
                    "must_generate": [
                        "case_data",
                        "application_data",
                        "review.missing_items",
                        "review.validation_errors",
                    ],
                },
                "applicant": {},
                "passport": {},
                "residence_card": {},
                "immigration_history": {},
                "family": {},
                "education": [],
                "transcript_subjects": [],
                "employment_history": [],
                "qualifications": [],
                "employer": {},
                "application": {},
                "supporting_documents": [],
                "assessments": [],
                "review": {
                    "status": "pending_manual_completion",
                    "notes": "このファイルを人手で正解データとして埋め、AI抽出結果との比較に使う。",
                },
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )
    (expected_dir / "application_data.golden.json").write_text("[]\n")
    (expected_dir / "review.golden.json").write_text(
        json.dumps(
            {
                "schema_version": "0.1.0",
                "case_id": case["case_id"],
                "golden_status": "pending_manual_completion",
                "expected_route": "needs_review",
                "expected_workflow_state": "needs_information",
                "missing_documents": [],
                "missing_items": [],
                "validation_errors": [],
                "findings": [],
                "assessments": [
                    {
                        "type": "gijinkoku_fit",
                        "status": "pending_manual_completion",
                        "summary": "人手で正解レビューを埋める。",
                    }
                ],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )
    (generated_dir / ".gitkeep").write_text("")


def main() -> None:
    if not RAW_ROOT.exists():
        raise SystemExit(f"missing raw root: {RAW_ROOT}")
    records = [record(path) for path in sorted(RAW_ROOT.rglob("*")) if path.is_file()]
    catalog = build_catalog(records)

    TEST_ROOT.mkdir(parents=True, exist_ok=True)
    (TEST_ROOT / "catalog.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n")
    for case in catalog["cases"]:
        write_fixture(case, records)
    print(f"wrote catalog for {len(catalog['cases'])} cases and {len(records)} documents")


if __name__ == "__main__":
    main()
