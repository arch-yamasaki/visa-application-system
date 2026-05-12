#!/usr/bin/env python3
"""Build restricted single-applicant golden fixtures from intake spreadsheets.

This is intentionally conservative: it uses structured Excel intake sheets as
the first source of truth and records PDFs/MSG/DOCX as supporting documents.
PDF OCR assertions should be added as separate extraction goldens later.
"""

from __future__ import annotations

import json
import re
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any

from dateutil import parser as date_parser
from openpyxl import load_workbook

from build_application_data import build_rows


ROOT = Path(__file__).resolve().parents[2]
TEST_ROOT = ROOT / "visa-eval"
CATALOG = TEST_ROOT / "catalog.json"
SINGLE_ROOT = TEST_ROOT / "fixtures_single"
MAPPING = ROOT / "rasens-autofill" / "data" / "mappings" / "rasens_offer_mapping.json"


def rel(path: Path) -> str:
    return str(path.relative_to(ROOT))


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def clean_question(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_answer(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = re.sub(r"\s+", " ", str(value)).strip()
    if text.endswith(" 00:00:00"):
        text = text[:-9]
    return text


def parse_date(value: str) -> str:
    if not value or value.upper() in {"NA", "N/A", "NO", "NONE"}:
        return ""
    try:
        return date_parser.parse(value, dayfirst=False, fuzzy=True).date().isoformat()
    except Exception:
        return value


def yes_no(value: str) -> bool | None:
    normalized = str(value or "").strip().lower()
    if normalized in {"yes", "y", "true", "有", "married"}:
        return True
    if normalized in {"no", "n", "false", "無", "na", "n/a", "none", "single", "unmarried"}:
        return False
    return None


def marital_status(value: str) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == "married":
        return "married"
    if normalized in {"single", "unmarried", "na", "n/a"}:
        return "single"
    return "unknown"


def question_key(question: str) -> str | None:
    q = question.lower()
    if "name (same as passport)" in q:
        return "name_roman"
    if "date of birth" in q and "その方" not in question:
        return "birth_date"
    if "place of birth" in q:
        return "birth_place"
    if "marital status" in q:
        return "marital_status"
    if "occupation" in q:
        return "occupation"
    if "hometown city" in q:
        return "home_country_address"
    if "passport number" in q:
        return "passport_number"
    if "passport date of expiration" in q:
        return "passport_expiry_date"
    if "past history of applying" in q:
        return "prior_coe_has_history"
    if "how many times did you apply for coe" in q:
        return "prior_coe_count"
    if "past entry into" in q:
        return "has_entries"
    if "how many times did you visit japan" in q:
        return "entries_count"
    if "latest entry to japan" in q:
        return "latest_entry_period_raw"
    if "criminal record" in q:
        return "criminal_record"
    if "do you have family in japan" in q:
        return "has_japan_relatives"
    if "relationship with him" in q:
        return "japan_relative_relationship"
    if "name of the family" in q:
        return "japan_relative_name"
    if "その方の生年月日" in question:
        return "japan_relative_birth_date"
    if "その方の国籍" in question:
        return "japan_relative_nationality"
    if "intended to reside" in q:
        return "japan_relative_live_together"
    if "name of the company/ school" in q:
        return "japan_relative_workplace_or_school"
    if "residence card number" in q:
        return "japan_relative_residence_card_number"
    if "name of last school" in q:
        return "school_name"
    if "date of graduation" in q:
        return "graduation_date"
    if "major field of study" in q:
        return "major"
    return None


def intake_pairs(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        qa: dict[str, Any] = {}
        raw_pairs = []
        for row in worksheet.iter_rows(min_row=1, max_row=300, values_only=True):
            values = list(row)
            candidates = []
            if len(values) > 4:
                candidates.append((values[3], values[4]))
            if len(values) > 1:
                candidates.append((values[0], values[1]))
            for question, answer in candidates:
                question_text = clean_question(question)
                answer_text = clean_answer(answer)
                if not question_text or question_text in {"Question", "@dropdown"} or answer_text == "":
                    continue
                key = question_key(question_text)
                if key:
                    qa[key] = answer_text
                raw_pairs.append({"question": question_text, "answer": answer_text, "key": key or "unmapped"})
                break
        sheets.append({"sheet_name": worksheet.title, "qa": qa, "raw_pairs": raw_pairs})
    workbook.close()
    return sheets


def split_period(raw: str) -> dict[str, str]:
    if not raw or raw.upper() in {"NA", "N/A", "NO"}:
        return {"start_date": "", "end_date": "", "raw": raw}
    parts = re.split(r"[~〜～-]+", raw)
    if len(parts) >= 2:
        return {"start_date": parse_date(parts[0]), "end_date": parse_date(parts[1]), "raw": raw}
    return {"start_date": "", "end_date": "", "raw": raw}


def applicant_from_sheet(sheet: dict[str, Any], source_path: str) -> dict[str, Any]:
    qa = sheet["qa"]
    name = qa.get("name_roman") or sheet["sheet_name"]
    has_relative = yes_no(qa.get("has_japan_relatives", ""))
    relative = {}
    if has_relative:
        relative = {
            "relationship": qa.get("japan_relative_relationship", ""),
            "full_name": qa.get("japan_relative_name", ""),
            "birth_date": parse_date(qa.get("japan_relative_birth_date", "")),
            "nationality_region": qa.get("japan_relative_nationality", ""),
            "will_live_together": yes_no(qa.get("japan_relative_live_together", "")),
            "workplace_or_school": qa.get("japan_relative_workplace_or_school", ""),
            "residence_card_number": qa.get("japan_relative_residence_card_number", ""),
        }

    return {
        "id": slug(name),
        "source_refs": [f"{source_path}#{sheet['sheet_name']}"],
        "applicant": {
            "name_roman": name,
            "birth_date": parse_date(qa.get("birth_date", "")),
            "birth_place": qa.get("birth_place", ""),
            "marital_status": marital_status(qa.get("marital_status", "")),
            "occupation": qa.get("occupation", ""),
            "home_country_address": qa.get("home_country_address", ""),
            "nationality_region": "ネパール Nepal",
            "sex": "unknown",
        },
        "passport": {
            "number": qa.get("passport_number", ""),
            "expiry_date": parse_date(qa.get("passport_expiry_date", "")),
        },
        "immigration_history": {
            "has_entries": yes_no(qa.get("has_entries", "")),
            "entries_count": qa.get("entries_count", ""),
            "latest_entry": split_period(qa.get("latest_entry_period_raw", "")),
            "prior_coe_applications": {
                "has_history": yes_no(qa.get("prior_coe_has_history", "")),
                "count": qa.get("prior_coe_count", ""),
                "denial_count": "",
            },
            "criminal_record": yes_no(qa.get("criminal_record", "")),
            "deportation_or_departure_order": None,
        },
        "family": {
            "has_accompanying_members": None,
            "has_japan_relatives_or_cohabitants": bool(has_relative),
            "japan_relatives_or_cohabitants": [relative] if has_relative else [],
        },
        "education": [
            {
                "id": "edu_01",
                "school_name": qa.get("school_name", ""),
                "graduation_date": parse_date(qa.get("graduation_date", "")),
                "major": qa.get("major", ""),
                "source_refs": [f"{source_path}#{sheet['sheet_name']}"],
            }
        ],
        "employment_history": [],
        "qualifications": [],
        "raw_intake_pairs": sheet["raw_pairs"],
    }


def load_catalog() -> dict[str, Any]:
    return json.loads(CATALOG.read_text())


def case_documents(catalog: dict[str, Any], case_id: str) -> list[dict[str, Any]]:
    return [doc for doc in catalog["documents"] if doc["case_id"] == case_id]


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def pick_documents(all_docs: list[dict[str, Any]], applicant_name: str) -> list[dict[str, Any]]:
    docs = []
    normalized = normalize_name(applicant_name)
    for doc in all_docs:
        if doc["document_role"] in {
            "intake_spreadsheet",
            "company_documents",
            "submitted_application_bundle",
            "not_attached_reference",
            "email_context",
        }:
            docs.append(doc)
            continue
        if normalized and normalize_name(Path(doc["file_name"]).stem).find(normalized) >= 0:
            docs.append(doc)
            continue
        if doc["document_role"] == "employment_terms" and normalized and normalize_name(doc["file_name"]).find(normalized) >= 0:
            docs.append(doc)
    if not any(d["document_role"] == "applicant_document_bundle" for d in docs):
        docs.extend([doc for doc in all_docs if doc["document_role"] == "applicant_document_bundle"])
    unique = {}
    for doc in docs:
        unique[doc["path"]] = doc
    return list(unique.values())


def intake_file(catalog: dict[str, Any], case_id: str) -> Path:
    docs = [doc for doc in case_documents(catalog, case_id) if doc["document_role"] == "intake_spreadsheet"]
    if not docs:
        raise SystemExit(f"no intake spreadsheet for {case_id}")
    return ROOT / docs[0]["path"]


def build_case(case: dict[str, Any], catalog: dict[str, Any]) -> dict[str, Any]:
    path = intake_file(catalog, case["case_id"])
    source = rel(path)
    applicants = [applicant_from_sheet(sheet, source) for sheet in intake_pairs(path)]
    documents = case_documents(catalog, case["case_id"])

    missing_items = []
    if any(app["applicant"]["sex"] == "unknown" for app in applicants):
        missing_items.append({"path": "applicants[].applicant.sex", "reason": "intake_spreadsheet_does_not_include_sex"})
    if any(app["immigration_history"]["deportation_or_departure_order"] is None for app in applicants):
        missing_items.append({"path": "applicants[].immigration_history.deportation_or_departure_order", "reason": "not_in_intake_spreadsheet"})

    return {
        "schema_version": "1.0",
        "golden_status": "restricted_real_data_from_intake_spreadsheet",
        "case": {
            "case_id": case["case_id"],
            "display_name": case["display_name"],
            "application_type": case["application_type"],
            "target_status": case["target_status"],
            "workflow_state": "golden_partial",
            "intake_channel": "recruiting_agency",
            "source_organization": "A社",
            "routed_to_human_reason": [],
        },
        "application": {
            "desired_status_label": "技術・人文知識・国際業務",
            "activity_details": "",
            "activity_details_structured": {
                "duties": [],
                "simple_labor_risk_terms": [],
            },
        },
        "applicants": applicants,
        "employer": {
            "name": "株式会社フジタ",
            "category": "unknown",
            "source_refs": [doc["path"] for doc in documents if doc["document_role"] == "company_documents"],
        },
        "supporting_documents": [
            {
                "document_id": f"doc_{index + 1:03d}",
                "document_type": doc["document_role"],
                "path": doc["path"],
                "file_name": doc["file_name"],
                "status": "received" if doc["use_as_input"] else "reference_only",
                "use_as_input": doc["use_as_input"],
                "contains_personal_information": doc["contains_personal_information"],
                "source_channel": "download",
            }
            for index, doc in enumerate(documents)
        ],
        "assessments": [
            {
                "type": "gijinkoku_fit",
                "status": "needs_review",
                "summary": "Structured intake was parsed. Education/job fit and activity details require review against PDFs and employer documents.",
            }
        ],
        "review": {
            "status": "needs_review",
            "missing_items": missing_items,
            "validation_errors": [],
            "notes": "Golden is populated from intake spreadsheet. PDF bundle and company documents still need human/OCR verification.",
        },
    }


def build_review(case_data: dict[str, Any]) -> dict[str, Any]:
    applicants = case_data.get("applicants", [])
    return {
        "schema_version": "0.1.0",
        "case_id": case_data["case"]["case_id"],
        "golden_status": case_data["golden_status"],
        "expected_route": "needs_review",
        "expected_workflow_state": "golden_partial",
        "missing_documents": [],
        "missing_items": case_data["review"]["missing_items"],
        "validation_errors": [],
        "findings": [
            {
                "code": "pdf_bundle_not_fully_verified",
                "severity": "medium",
                "message": "Applicant PDFs and company documents are linked as inputs but not fully normalized in golden yet.",
            }
        ],
        "assessments": case_data["assessments"],
        "stats": {
            "applicant_count": len(applicants),
            "has_japan_relatives_count": sum(
                1 for applicant in applicants if applicant.get("family", {}).get("has_japan_relatives_or_cohabitants")
            ),
        },
    }


def single_case_from_intake(case_data: dict[str, Any], applicant: dict[str, Any], docs: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": case_data["schema_version"],
        "golden_status": "restricted_single_applicant_fixture",
        "case": {
            "case_id": f"{case_data['case']['case_id']}__{applicant['id']}",
            "display_name": f"{case_data['case']['display_name']} / {applicant['applicant']['name_roman']}",
            "application_type": case_data["case"]["application_type"],
            "target_status": case_data["case"]["target_status"],
            "workflow_state": case_data["case"]["workflow_state"],
            "intake_channel": case_data["case"]["intake_channel"],
            "source_organization": case_data["case"]["source_organization"],
            "routed_to_human_reason": [],
        },
        "application": deepcopy(case_data["application"]),
        "applicant": deepcopy(applicant["applicant"]),
        "passport": deepcopy(applicant["passport"]),
        "residence_card": {},
        "immigration_history": deepcopy(applicant["immigration_history"]),
        "family": deepcopy(applicant["family"]),
        "education": deepcopy(applicant["education"]),
        "transcript_subjects": [],
        "employment_history": deepcopy(applicant["employment_history"]),
        "qualifications": deepcopy(applicant["qualifications"]),
        "employer": deepcopy(case_data["employer"]),
        "proxy": {},
        "intermediary": {},
        "receiving_method": {},
        "supporting_documents": [
            {
                "document_id": f"doc_{index + 1:03d}",
                "document_type": doc["document_role"],
                "path": doc["path"],
                "file_name": doc["file_name"],
                "status": "received" if doc["use_as_input"] else "reference_only",
                "use_as_input": doc["use_as_input"],
                "contains_personal_information": doc["contains_personal_information"],
                "source_channel": "download",
            }
            for index, doc in enumerate(docs)
        ],
        "assessments": deepcopy(case_data["assessments"]),
        "field_metadata": {},
    }


def review_from_single_case(case_data: dict[str, Any]) -> dict[str, Any]:
    missing_items = []
    if case_data.get("applicant", {}).get("sex") in {"", "unknown"}:
        missing_items.append({"path": "applicant.sex", "reason": "intake_spreadsheet_does_not_include_sex"})
    if case_data.get("immigration_history", {}).get("deportation_or_departure_order") is None:
        missing_items.append({"path": "immigration_history.deportation_or_departure_order", "reason": "not_in_intake_spreadsheet"})
    if not case_data.get("residence_card"):
        missing_items.append({"path": "residence_card", "reason": "not_present_in_intake_spreadsheet"})

    return {
        "schema_version": "0.1.0",
        "case_id": case_data["case"]["case_id"],
        "golden_status": case_data["golden_status"],
        "expected_route": "needs_review",
        "expected_workflow_state": case_data["case"]["workflow_state"],
        "missing_documents": [],
        "missing_items": missing_items,
        "validation_errors": [],
        "findings": [
            {
                "code": "pdf_bundle_not_fully_verified",
                "severity": "medium",
                "message": "Intake spreadsheet was parsed; applicant PDFs and company documents still need human/OCR verification.",
            }
        ],
        "assessments": deepcopy(case_data["assessments"]),
    }


def write_single_fixture(case_data: dict[str, Any], docs: list[dict[str, Any]], mapping: dict[str, Any]) -> None:
    case_id, applicant_id = case_data["case"]["case_id"].split("__", 1)
    fixture_dir = SINGLE_ROOT / case_id / applicant_id
    input_dir = fixture_dir / "input"
    expected_dir = fixture_dir / "expected"
    generated_dir = fixture_dir / "generated"
    input_dir.mkdir(parents=True, exist_ok=True)
    expected_dir.mkdir(parents=True, exist_ok=True)
    generated_dir.mkdir(parents=True, exist_ok=True)

    (fixture_dir / "scenario.json").write_text(
        json.dumps(
            {
                "schema_version": "0.1.0",
                "case_id": case_data["case"]["case_id"],
                "base_case_id": case_id,
                "applicant_id": applicant_id,
                "applicant_name": case_data["applicant"]["name_roman"],
                "target_status": case_data["case"]["target_status"],
                "application_type": case_data["case"]["application_type"],
                "primary_checks": [
                    "document_completeness",
                    "case_data_generation",
                    "form_mapping",
                    "gijinkoku_fit",
                ],
                "input_documents": "input/input_documents.json",
                "expected_case_data": "expected/case_data.golden.json",
                "expected_application_data": "expected/application_data.golden.json",
                "expected_review": "expected/review.golden.json",
                "generated_dir": "generated",
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )

    (input_dir / "input_documents.json").write_text(
        json.dumps(
            {
                "schema_version": "0.1.0",
                "case_id": case_data["case"]["case_id"],
                "base_case_id": case_id,
                "applicant_id": applicant_id,
                "documents": docs,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )

    (expected_dir / "case_data.golden.json").write_text(json.dumps(case_data, ensure_ascii=False, indent=2) + "\n")
    (expected_dir / "review.golden.json").write_text(json.dumps(review_from_single_case(case_data), ensure_ascii=False, indent=2) + "\n")
    (expected_dir / "application_data.golden.json").write_text(json.dumps(build_rows(case_data, mapping), ensure_ascii=False, indent=2) + "\n")
    (generated_dir / ".gitkeep").write_text("")


def main() -> None:
    catalog = load_catalog()
    mapping = json.loads(MAPPING.read_text())
    for case in catalog["cases"]:
        case_data = build_case(case, catalog)
        all_docs = case_documents(catalog, case["case_id"])
        for applicant in case_data["applicants"]:
            docs = pick_documents(all_docs, applicant["applicant"]["name_roman"])
            single = single_case_from_intake(case_data, applicant, docs)
            write_single_fixture(single, docs, mapping)
            print(f"wrote {single['case']['case_id']}")


if __name__ == "__main__":
    main()
