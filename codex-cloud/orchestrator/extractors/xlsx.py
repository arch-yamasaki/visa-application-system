"""openpyxl によるExcelテキスト抽出。"""

import io

import openpyxl

from .types import OcrResult, PageResult


def extract_xlsx(file_bytes: bytes, document_id: str, sheet_name: str | None = None) -> OcrResult:
    """Excelからテキストを抽出。シート=ページとして扱う。"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets
    pages = []
    for i, ws in enumerate(sheets):
        lines = []
        for row in ws.iter_rows(values_only=False):
            cells = []
            for cell in row:
                if cell.value is not None:
                    cells.append(str(cell.value))
            if cells:
                lines.append("\t".join(cells))
        text = "\n".join(lines)
        if text.strip():
            pages.append(PageResult(page_number=i + 1, text=f"[Sheet: {ws.title}]\n{text}"))
    wb.close()
    return OcrResult(document_id=document_id, pages=pages)
